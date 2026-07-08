importimport io
import os
import re
from datetime importimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
fromimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context importimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
fromimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxlimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import getimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
fromimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", autoimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: strimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = Noneimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expiresimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expireimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(toimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(tokenimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_schemeimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raiseimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwtimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db:import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    ifimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    tokenimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.adminimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": tokenimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/meimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin))import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.getimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.Statsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(getimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).countimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=dbimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        totalimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOutimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optionalimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page:import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filterimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{qimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) *import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * pageimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categoriesimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c inimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_ormimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.getimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}",import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resourceimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Dependsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin))import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raiseimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).firstimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = categoryimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return dataimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources",import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resourceimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session =import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_adminimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payloadimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filterimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_ormimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.nameimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resourceimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id:import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: intimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filterimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPExceptionimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dictimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    dbimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_idimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOutimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_nameimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resourceimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filterimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPExceptionimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    dbimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commitimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}

import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def importimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile =import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Dependsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_currentimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filenameimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsximport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls"import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIOimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contentsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wbimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rowsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categoriesimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c inimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail =import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support bothimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row inimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not rowimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).stripimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).stripimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for vimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool ximport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        ifimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            contentimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = rowimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        #import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ...,import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            contentimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = rowimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(modelsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resourceimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if existsimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continueimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_nameimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parseimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.nowimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=contentimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            createdimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success +=import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}，失败条数：{fail}"}
import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}，失败条数：{fail}"}


def parse_time(value) -> Optional[import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}，失败条数：{fail}"}


def parse_time(value) -> Optional[datetime]:
    if isinstance(value, datetimeimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}，失败条数：{fail}"}


def parse_time(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if not valueimport io
import os
import re
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.config import get_settings

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login", auto_error=False)

settings = get_settings()


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.access_token_expire_minutes))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm="HS256")


async def get_current_admin(token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=["HS256"])
        username: str = payload.get("sub")
        if username != settings.admin_username:
            raise HTTPException(status_code=401, detail="无效凭证")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效凭证")
    return {"username": username}


@router.post("/login", response_model=schemas.TokenOut)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    if form_data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if not verify_password(form_data.password, settings.admin_password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": settings.admin_username})
    return {"access_token": token, "token_type": "bearer"}


@router.get("/me")
def me(admin=Depends(get_current_admin)):
    return admin


@router.get("/stats", response_model=schemas.StatsOut)
def stats(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return schemas.StatsOut(
        total_resources=db.query(models.Resource).count(),
        total_categories=db.query(models.Category).count(),
        total_keywords=db.query(models.Keyword).count(),
        total_reports=db.query(models.Report).count(),
    )


@router.get("/resources", response_model=list[schemas.ResourceOut])
def list_resources(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    query = db.query(models.Resource)
    if q:
        query = query.filter(models.Resource.title.ilike(f"%{q}%"))
    resources = (
        query.order_by(desc(models.Resource.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    categories = {c.id: c.name for c in db.query(models.Category).all()}
    out = []
    for r in resources:
        item = schemas.ResourceOut.from_orm(r)
        item.category_name = categories.get(r.category_id, "")
        out.append(item)
    return out


@router.get("/resources/{resource_id}", response_model=schemas.ResourceOut)
def get_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.post("/resources", response_model=schemas.ResourceOut)
def create_resource(payload: schemas.ResourceCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = models.Resource(**payload.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.put("/resources/{resource_id}", response_model=schemas.ResourceOut)
def update_resource(
    resource_id: int,
    payload: schemas.ResourceUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    for k, v in payload.dict().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    category = db.query(models.Category).filter(models.Category.id == r.category_id).first()
    data = schemas.ResourceOut.from_orm(r)
    data.category_name = category.name if category else ""
    return data


@router.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="资源不存在")
    db.delete(r)
    db.commit()
    return {"status": 1, "message": "删除成功"}


@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件类型")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入数据为空")

    categories = {c.name: c.id for c in db.query(models.Category).all()}
    success = 0
    fail = 0

    # Skip header row, support both pantool and generic formats
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]

        # Pantool xls format: title, content, url, code, category_name, size, time
        if len(row) >= 7 and row[4] and row[4] in categories:
            title = row[0]
            content = row[1]
            url = row[2]
            code = row[3]
            category_name = row[4]
            size = row[5]
            time_str = row[6]
        # Generic xlsx format: id, title, ..., url, code, category_name, ..., size
        elif len(row) >= 13:
            title = row[1]
            content = row[1]
            url = row[3]
            code = row[4]
            category_name = row[5]
            size = row[12]
            time_str = ""
        else:
            fail += 1
            continue

        if not title or not url:
            fail += 1
            continue

        exists = db.query(models.Resource).filter(models.Resource.title == title).first()
        if exists:
            fail += 1
            continue

        category_id = categories.get(category_name, 0)
        created_at = parse_time(time_str) if time_str else datetime.now()

        r = models.Resource(
            title=title,
            content=content,
            url=url,
            code=code,
            size=size,
            category_id=category_id,
            created_at=created_at,
        )
        db.add(r)
        success += 1

    db.commit()
    return {"status": 1, "message": f"导入成功，成功条数：{success}，失败条数：{fail}"}


def parse_time(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    text = str